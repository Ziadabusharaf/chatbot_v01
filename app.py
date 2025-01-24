import os
from flask import Flask, request, render_template, jsonify
from openpyxl import load_workbook
from PyPDF2 import PdfReader
import openai

# Configure OpenAI API Key (replace with your actual key)
openai.api_key = "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"

app = Flask(__name__)

# Configure upload folder
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


@app.route('/')
def home():
    """Render the main chatbot and upload interface."""
    return render_template('index.html')


@app.route('/chat', methods=['POST'])
def chat():
    """Handle user input and provide AI-generated responses."""
    user_message = request.json.get('message', '')
    response = generate_response(user_message)
    return jsonify({"response": response})


@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file uploads and return results in chat response."""
    if 'file' not in request.files:
        return jsonify({"response": "No file found. Please attach a file to upload."}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"response": "No file selected. Please choose a file to upload."}), 400

    # Save the file to the uploads folder
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(file_path)

    # Process Excel files
    if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
        result = process_excel(file_path)
        return jsonify({
            "response": f"Excel file '{file.filename}' processed successfully!",
            "data": result["data"],
            "validations": result["validation_results"]
        })
    elif file.filename.endswith('.pdf'):
        text = process_pdf(file_path)
        return jsonify({
            "response": f"PDF file '{file.filename}' processed successfully!",
            "text": text
        })
    else:
        return jsonify({"response": "Unsupported file format. Please upload an Excel or PDF file."})


def generate_response(user_message):
    """AI-driven response generator using OpenAI GPT."""
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",  # Use "gpt-4" if available
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": user_message}
            ],
            max_tokens=150,
            temperature=0.7,
            stream=True  # Stream the response
        )
        streamed_response = ""
        for chunk in response:
            if "choices" in chunk and chunk["choices"][0]["delta"].get("content"):
                streamed_response += chunk["choices"][0]["delta"]["content"]
        return streamed_response
    except openai.error.AuthenticationError:
        return "Error: Invalid API key. Please check your OpenAI API key."
    except openai.error.RateLimitError:
        return "Error: Rate limit exceeded. Please wait and try again."
    except openai.error.APIError as api_err:
        return f"Error: API error occurred: {str(api_err)}"
    except openai.error.OpenAIError as e:
        return f"Error: An unexpected OpenAI error occurred: {str(e)}"
    except Exception as e:
        return f"Error: An unexpected error occurred: {str(e)}"


def process_excel(file_path):
    """Process the uploaded Excel file and validate data."""
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    validation_results = []

    for row in sheet.iter_rows(values_only=True):
        if row[0] == "ID":
            data.append(list(row) + ["Validation"])
            continue
        row_data = list(row)
        validation = validate_engineering_rules(row)
        row_data.append(validation)
        data.append(row_data)
        validation_results.append(validation)

    return {"data": data, "validation_results": validation_results}


def validate_engineering_rules(row):
    """Validate engineering rules."""
    try:
        overhang = row[3]  # Assume overhang is in the 4th column
        if 0.5 <= overhang <= 3.0:
            return "Pass"
        else:
            return f"Fail: Overhang {overhang} out of bounds (0.5-3.0)"
    except (IndexError, TypeError):
        return "Fail: Missing or invalid data"


def process_pdf(file_path):
    """Extract text from the uploaded PDF file."""
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"
    return text


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=8080, debug=True)
import sqlite3

def save_conversation(user_message, bot_response):
    # Connect to SQLite database
    conn = sqlite3.connect("chatbot.db")
    cursor = conn.cursor()

    # Insert the conversation
    cursor.execute("""
        INSERT INTO conversations (user_message, bot_response)
        VALUES (?, ?)
    """, (user_message, bot_response))

    # Commit changes and close connection
    conn.commit()
    conn.close()

# Example usage
save_conversation("Hello!", "Hi there!")
